"""
灵动智能体 - 可编辑分镜表格节点
支持：
1. 直接在画布上编辑（无需弹窗）
2. 可编辑表格名称
3. 可编辑表头
4. 可编辑单元格内容
5. 选中行高亮（红色半透明）
6. 可自由缩放调整大小
7. 右键单元格：复制、修改内容
8. 长文本自动省略显示（超过80字符显示前75字符+"..."）
9. 鼠标悬停查看完整内容（工具提示）
10. 双击单元格进入编辑模式
11. 编辑状态下显示完整文本
12. 右上角折叠/展开按钮（▼/▶），点击可隐藏/显示表格内容
13. 绘画提示词列（CN/EN）支持垂直滚动，文字不会被压缩或溢出
"""

from PySide6.QtWidgets import (QGraphicsRectItem, QGraphicsTextItem, QGraphicsLineItem,
                                QMenu, QGraphicsItem, QInputDialog, QLineEdit, QFileDialog, QMessageBox,
                                QGraphicsProxyWidget, QTextEdit, QScrollBar, QGraphicsPixmapItem,
                                QDialog, QVBoxLayout, QLabel, QPlainTextEdit, QDialogButtonBox)
from PySide6.QtCore import Qt, QRectF, Signal, QObject, QSize
from PySide6.QtGui import QFont, QColor, QPen, QBrush, QAction, QTextCursor, QPalette, QPainter, QPixmap, QPainterPath
import os


# SVG图标定义
SVG_TABLE_ICON = '''<svg viewBox="0 0 24 24" fill="none" xmlns="http://www.w3.org/2000/svg">
<rect x="3" y="3" width="18" height="18" rx="2" stroke="currentColor" stroke-width="2"/>
<path d="M3 9h18M3 15h18M9 3v18" stroke="currentColor" stroke-width="2"/>
</svg>'''


class EditableTextItem(QGraphicsTextItem):
    """可编辑的文本项 - 支持长文本滚动和工具提示"""
    
    def __init__(self, text, parent=None, is_header=False, max_lines=2):
        super().__init__(text, parent)
        self.is_header = is_header
        self.full_text = text  # 保存完整文本
        self.max_lines = max_lines  # 最大显示行数
        self.is_editing = False  # 是否正在编辑
        self.is_expanded = False  # 是否展开显示
        
        self.setTextInteractionFlags(Qt.TextInteractionFlag.TextEditorInteraction)
        self.setDefaultTextColor(QColor("#1a73e8" if is_header else "#202124"))
        self.setFont(QFont("Microsoft YaHei", 9 if is_header else 8, 
                          QFont.Weight.Bold if is_header else QFont.Weight.Normal))
        
        # 设置为可选择和可聚焦
        self.setFlag(QGraphicsItem.GraphicsItemFlag.ItemIsFocusable)
        self.setFlag(QGraphicsItem.GraphicsItemFlag.ItemIsSelectable)
        
        # 启用工具提示
        self.setAcceptHoverEvents(True)
        if len(text) > 50:
            self.setToolTip(f"完整内容:\n{text}\n\n💡 双击可编辑")
        else:
            self.setToolTip("💡 双击可编辑")
    
    def contextMenuEvent(self, event):
        """禁用默认右键菜单（undo, redo, cut, copy, paste等）"""
        event.ignore()  # 忽略事件，让父节点处理
    
    def focusInEvent(self, event):
        """获得焦点时选中所有文本并显示完整内容"""
        self.is_editing = True
        # 编辑时显示完整文本
        self.setPlainText(self.full_text)
        super().focusInEvent(event)
        cursor = self.textCursor()
        cursor.select(QTextCursor.SelectionType.Document)
        self.setTextCursor(cursor)
    
    def focusOutEvent(self, event):
        """失去焦点时保存内容并恢复省略显示"""
        self.is_editing = False
        # 保存用户编辑的内容
        self.full_text = self.toPlainText()
        
        # 更新工具提示
        if len(self.full_text) > 50:
            self.setToolTip(f"完整内容:\n{self.full_text}\n\n💡 双击可编辑")
        else:
            self.setToolTip("💡 双击可编辑")
        
        super().focusOutEvent(event)
        # 恢复省略显示（如果内容超出）
        self.update_display()
    
    def hoverEnterEvent(self, event):
        """鼠标悬停时高亮显示"""
        if not self.is_editing:
            # 改变颜色提示可点击
            self.setDefaultTextColor(QColor("#174ea6" if self.is_header else "#000000"))
        super().hoverEnterEvent(event)
    
    def hoverLeaveEvent(self, event):
        """鼠标离开时恢复颜色"""
        if not self.is_editing:
            self.setDefaultTextColor(QColor("#1a73e8" if self.is_header else "#202124"))
        super().hoverLeaveEvent(event)

    def mousePressEvent(self, event):
        if event.button() == Qt.MouseButton.LeftButton:
            parent = self.parentItem()
            if parent and hasattr(parent, "select_cell"):
                row_idx = self.data(0)
                col_idx = self.data(1)
                if row_idx is not None and col_idx is not None:
                    try:
                        parent.select_cell(int(row_idx), int(col_idx) + 1)
                    except Exception:
                        pass
        super().mousePressEvent(event)
    
    def mouseDoubleClickEvent(self, event):
        """双击进入编辑模式"""
        # 设置为可编辑并获得焦点
        self.setFocus()
        super().mouseDoubleClickEvent(event)
    
    def update_display(self):
        """更新显示内容（编辑外显示省略，编辑时显示完整）"""
        if not self.is_editing and not self.is_expanded:
            # 非编辑状态，检查是否需要省略
            # 简单策略：如果文本长度超过80字符，显示前75字符+"..."
            if len(self.full_text) > 80:
                display_text = self.full_text[:75] + "..."
                self.setPlainText(display_text)
            else:
                self.setPlainText(self.full_text)
        elif self.is_expanded:
            # 展开状态显示完整内容
            self.setPlainText(self.full_text)


class ScrollableTextCell(QGraphicsProxyWidget):
    """可滚动的文本单元格 - 使用QTextEdit实现真正的滚动"""
    
    def __init__(self, text, parent, width, height, is_header=False):
        super().__init__(parent)
        
        self.is_header = is_header
        self.cell_width = width
        self.cell_height = height
        
        # 创建QTextEdit
        self.text_edit = QTextEdit()
        self.text_edit.setPlainText(text)
        
        # 禁用QTextEdit的默认右键菜单
        self.text_edit.setContextMenuPolicy(Qt.ContextMenuPolicy.NoContextMenu)
        
        # 设置字体
        font = QFont("Microsoft YaHei", 9 if is_header else 8, 
                    QFont.Weight.Bold if is_header else QFont.Weight.Normal)
        self.text_edit.setFont(font)
        
        # 设置样式
        self.text_edit.setStyleSheet(f"""
            QTextEdit {{
                background-color: transparent;
                color: {'#188038' if is_header else '#333333'};
                border: none;
                padding: 2px;
            }}
            QTextEdit:focus {{
                background-color: #ffffff;
                color: {'#188038' if is_header else '#000000'};
                border: 1px solid #34A853;
            }}
            QScrollBar:vertical {{
                background: #f5f5f5;
                width: 8px;
                border-radius: 4px;
            }}
            QScrollBar::handle:vertical {{
                background: #bdbdbd;
                border-radius: 4px;
                min-height: 20px;
            }}
            QScrollBar::handle:vertical:hover {{
                background: #a0a0a0;
            }}
            QScrollBar::add-line:vertical, QScrollBar::sub-line:vertical {{
                height: 0px;
            }}
        """)
        
        # 设置大小
        self.text_edit.setFixedSize(int(width - 10), int(height - 4))
        
        # 启用自动换行
        self.text_edit.setLineWrapMode(QTextEdit.LineWrapMode.WidgetWidth)
        
        # 垂直滚动条始终显示
        self.text_edit.setVerticalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAsNeeded)
        self.text_edit.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOff)
        
        # 设置为代理控件
        self.setWidget(self.text_edit)
        
        # 工具提示
        if len(text) > 50:
            self.setToolTip(f"完整内容:\n{text}\n\n💡 可直接编辑和滚动")
        else:
            self.setToolTip("💡 可直接编辑")
    
    def toPlainText(self):
        """获取纯文本（兼容接口）"""
        return self.text_edit.toPlainText()
    
    def setPlainText(self, text):
        """设置纯文本（兼容接口）"""
        self.text_edit.setPlainText(text)
    
    def contextMenuEvent(self, event):
        """禁用默认右键菜单，让父节点处理"""
        event.ignore()  # 忽略事件，让父节点处理

    def mousePressEvent(self, event):
        if event.button() == Qt.MouseButton.LeftButton:
            parent = self.parentItem()
            if parent and hasattr(parent, "select_cell"):
                row_idx = self.data(0)
                col_idx = self.data(1)
                if row_idx is not None and col_idx is not None:
                    try:
                        parent.select_cell(int(row_idx), int(col_idx) + 1)
                    except Exception:
                        pass
        super().mousePressEvent(event)


class ImageCell(QGraphicsRectItem):
    """图片单元格 - 用于显示缩略图"""
    
    def __init__(self, image_path, width, height, parent=None):
        super().__init__(0, 0, width, height, parent)
        self.image_path = image_path
        self.width = width
        self.height = height
        
        # 设置背景透明
        self.setBrush(Qt.BrushStyle.NoBrush)
        self.setPen(Qt.PenStyle.NoPen)
        
        # 如果有图片路径，加载图片
        if image_path and os.path.exists(image_path):
            try:
                pixmap = QPixmap(image_path)
                if not pixmap.isNull():
                    # 计算缩放比例，保持纵横比适应单元格高度
                    # 留出一点边距
                    target_h = height - 4
                    scaled_pixmap = pixmap.scaledToHeight(int(target_h), Qt.TransformationMode.SmoothTransformation)
                    
                    # 如果宽度超过单元格宽度，则按宽度缩放
                    if scaled_pixmap.width() > width - 4:
                        scaled_pixmap = pixmap.scaledToWidth(int(width - 4), Qt.TransformationMode.SmoothTransformation)
                    
                    # 创建图片项
                    self.pixmap_item = QGraphicsPixmapItem(scaled_pixmap, self)
                    
                    # 居中显示
                    x_pos = (width - scaled_pixmap.width()) / 2
                    y_pos = (height - scaled_pixmap.height()) / 2
                    self.pixmap_item.setPos(x_pos, y_pos)
                    
                    # 设置工具提示
                    self.setToolTip(f"点击查看大图\n{image_path}")
                    
                    # 设置鼠标指针为手型
                    self.setCursor(Qt.CursorShape.PointingHandCursor)
                    
            except Exception as e:
                print(f"[图片单元格] 加载图片失败: {e}")
                self.show_text(image_path)
        else:
            self.show_text(image_path if image_path else "")
        
        # 允许鼠标交互
        self.setAcceptedMouseButtons(Qt.MouseButton.LeftButton)

    def show_text(self, text):
        """显示文本（当没有图片或加载失败时）"""
        text_item = QGraphicsTextItem(text, self)
        text_item.setDefaultTextColor(QColor("#5f6368"))
        text_item.setFont(QFont("Microsoft YaHei", 8))
        
        # 裁剪过长文本
        if len(text) > 15:
             text_item.setPlainText(text[:12] + "...")
             text_item.setToolTip(text)
        
        # 居中
        text_rect = text_item.boundingRect()
        x_pos = (self.width - text_rect.width()) / 2
        y_pos = (self.height - text_rect.height()) / 2
        text_item.setPos(x_pos, y_pos)

    def mousePressEvent(self, event):
        """鼠标点击事件 - 单击选中并打开图片"""
        if event.button() == Qt.MouseButton.LeftButton:
            # 1. 调用父节点的选中逻辑
            if self.parentItem() and hasattr(self.parentItem(), 'select_cell'):
                row_idx = self.data(0)
                col_idx = self.data(1)
                
                if row_idx is not None and col_idx is not None:
                    try:
                        # 转换列索引：数据列索引 -> 表头索引
                        header_idx = int(col_idx) + 1
                        self.parentItem().select_cell(int(row_idx), header_idx)
                            
                    except Exception as e:
                        # print(f"[图片单元格] 选中失败: {e}")
                        pass
            
            # 2. 打开图片
            if self.image_path and os.path.exists(self.image_path):
                try:
                    os.startfile(self.image_path)
                except Exception as e:
                    print(f"[图片单元格] 打开图片失败: {e}")
            
            event.accept()
        else:
            super().mousePressEvent(event)

    def mouseDoubleClickEvent(self, event):
        """双击事件"""
        # 由于单击已经打开了图片，双击不再处理，或者是作为备份
        super().mouseDoubleClickEvent(event)


class CellEditDialog(QDialog):
    """自定义单元格编辑对话框 - 优化大段文字编辑体验"""
    def __init__(self, title, label_text, content, parent=None):
        super().__init__(parent)
        self.setWindowTitle(title)
        self.setMinimumSize(900, 600)  # 设置较大的初始尺寸
        self.resize(1000, 700) # 默认宽一点，方便显示长文本
        
        # 设置样式
        self.setStyleSheet("""
            QDialog {
                background-color: #ffffff;
                color: #333333;
            }
            QLabel {
                color: #34A853;
                font-size: 14px;
                font-weight: bold;
                margin-bottom: 10px;
            }
            QPlainTextEdit {
                background-color: #f8f9fa;
                color: #333333;
                border: 1px solid #e0e0e0;
                border-radius: 4px;
                padding: 10px;
                font-size: 14px;
                line-height: 1.5;
            }
            QButtonBox {
                margin-top: 10px;
            }
            QPushButton {
                background-color: #ffffff;
                color: #333333;
                border: 1px solid #e0e0e0;
                border-radius: 4px;
                padding: 6px 15px;
                min-width: 80px;
            }
            QPushButton:hover {
                background-color: #f1f3f4;
                border: 1px solid #d0d0d0;
            }
            QPushButton:pressed {
                background-color: #e6f4ea;
            }
        """)
        
        layout = QVBoxLayout(self)
        layout.setContentsMargins(20, 20, 20, 20)
        
        # 标签
        label = QLabel(label_text)
        layout.addWidget(label)
        
        # 文本编辑器
        self.text_edit = QPlainTextEdit()
        self.text_edit.setPlainText(content)
        self.text_edit.setLineWrapMode(QPlainTextEdit.LineWrapMode.WidgetWidth) # 自动换行
        
        # 设置字体
        font = QFont("Microsoft YaHei", 12)
        self.text_edit.setFont(font)
        
        layout.addWidget(self.text_edit)
        
        # 按钮
        buttons = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        buttons.accepted.connect(self.accept)
        buttons.rejected.connect(self.reject)
        layout.addWidget(buttons)
        
    @property
    def text(self):
        return self.text_edit.toPlainText()


class ToggleButton(QGraphicsItem):
    """折叠/展开按钮"""
    def __init__(self, parent, callback):
        super().__init__(parent)
        self.callback = callback
        self.is_collapsed = False
        self.setAcceptHoverEvents(True)
        self.setAcceptedMouseButtons(Qt.MouseButton.LeftButton)
        self._rect = QRectF(0, 0, 24, 24)
        
    def boundingRect(self):
        return self._rect
        
    def paint(self, painter, option, widget):
        painter.setRenderHint(QPainter.RenderHint.Antialiasing)
        
        # 绘制图标
        painter.setPen(Qt.PenStyle.NoPen)
        painter.setBrush(QBrush(QColor("#5f6368")))
        
        if self.is_collapsed:
            # ▶ 折叠状态
            path = QPainterPath()
            path.moveTo(8, 6)
            path.lineTo(18, 12)
            path.lineTo(8, 18)
            path.closeSubpath()
            painter.drawPath(path)
        else:
            # ▼ 展开状态
            path = QPainterPath()
            path.moveTo(6, 8)
            path.lineTo(12, 18)
            path.lineTo(18, 8)
            path.closeSubpath()
            painter.drawPath(path)
            
    def update_icon(self, is_collapsed):
        self.is_collapsed = is_collapsed
        self.update()
        
    def mousePressEvent(self, event):
        if event.button() == Qt.MouseButton.LeftButton:
            if self.callback:
                self.callback()
            event.accept()
    
    def hoverEnterEvent(self, event):
        self.setCursor(Qt.CursorShape.PointingHandCursor)
        super().hoverEnterEvent(event)
        
    def hoverLeaveEvent(self, event):
        self.setCursor(Qt.CursorShape.ArrowCursor)
        super().hoverLeaveEvent(event)


class StoryboardNode:
    """可编辑分镜表格节点 - 直接在画布上编辑
    
    注意：这个类需要继承自CanvasNode，在导入时会动态继承
    """
    
    @staticmethod
    def create_storyboard_node(CanvasNode):
        """动态创建StoryboardNode类，继承自CanvasNode"""
        
        class StoryboardNodeImpl(CanvasNode):
            """可编辑分镜表格节点实现"""
            
            def __init__(self, x, y, table_data=None, table_name="分镜脚本", custom_headers=None):
                # 表格数据 - 如果没有提供数据，创建空表格（0行）
                self.table_data = table_data if table_data is not None else []
                self.original_table_name = table_name
                
                # 表格配置
                self.cell_height = 150  # 增加行高以显示更大的图片
                self.header_height = 40
                self.title_height = 45
                
                # 配置按钮位置（在标题栏右侧，折叠按钮左边）
                self.settings_btn_size = 20
                self.settings_btn_rect = None  # 将在paint中计算
                
                # ⭐ 支持自定义表头
                if custom_headers:
                    # 使用自定义表头
                    self.headers = custom_headers
                    # 根据表头数量动态计算列宽
                    num_cols = len(custom_headers)
                    if num_cols == 2:
                        # 2列：镜号 + 其他
                        self.column_widths = [60, 400]
                    elif num_cols == 3:
                        self.column_widths = [60, 300, 300]
                    elif num_cols <= 5:
                        base_width = 150
                        self.column_widths = [60] + [base_width] * (num_cols - 1)
                    else:
                        # 多列情况，第一列60，其他平均分配
                        base_width = 120
                        self.column_widths = [60] + [base_width] * (num_cols - 1)
                    
                    # 调整数据列数以匹配表头（去除镜号列）
                    expected_data_cols = num_cols - 1
                    if table_data:
                        self.table_data = []
                        for row in table_data:
                            if len(row) < expected_data_cols:
                                # 补齐空列
                                self.table_data.append(row + [""] * (expected_data_cols - len(row)))
                            else:
                                self.table_data.append(row[:expected_data_cols])
                    else:
                        # 创建空表格（0行）
                        self.table_data = []
                else:
                    # 使用默认10列配置（添加镜头号和时间码列）
                    self.column_widths = [80, 120, 100, 250, 120, 150, 120, 100, 250, 120]  # 增加列宽
                    self.headers = ["镜号", "时间码", "景别", "画面内容", "人物", "人物关系/构图", 
                                  "地点/环境", "运镜", "音效/台词", "备注"]
                
                # 计算初始大小
                total_width = sum(self.column_widths) + 2
                row_count = len(self.table_data)
                # 如果表格为空，保留一定的最小高度用于显示提示
                min_content_height = 80 if row_count == 0 else 0
                total_height = self.title_height + self.header_height + (row_count * self.cell_height) + min_content_height + 2
                
                # 初始化基类
                super().__init__(x, y, total_width, total_height, table_name, SVG_TABLE_ICON)
                
                # 隐藏默认标题
                if hasattr(self, 'title_text'):
                    self.title_text.setVisible(False)
                
                # 存储UI元素
                self.title_item = None
                self.header_items = []
                self.cell_items = []
                self.row_backgrounds = []
                self.selected_row = -1
                self.selected_col = -1  # 新增：选中的列索引
                self.selection_rect_item = None # 新增：选中区域的高亮矩形
                
                # 缩放控制
                self.is_resizing = False
                self.resize_start_pos = None
                self.resize_handle_size = 15
                
                # 折叠/展开状态
                self.is_collapsed = False  # 默认展开
                self.toggle_button = None  # 折叠按钮
                self.collapsed_height = self.title_height + 2  # 折叠后的高度
                self.expanded_height = 0  # 展开后的高度（动态计算）
                
                # 创建表格
                self.create_table()
                
                # 设置为可调整大小
                self.setFlag(QGraphicsItem.GraphicsItemFlag.ItemIsSelectable)
                self.setFlag(QGraphicsItem.GraphicsItemFlag.ItemIsMovable)
            
            def _set_content_visible(self, visible):
                """设置内容可见性 - 覆盖基类方法以保留标题"""
                # 调用基类方法处理大部分元素
                super()._set_content_visible(visible)
                
                # 确保自定义标题始终可见
                if self.title_item:
                    self.title_item.setVisible(True)
                
                # 确保选中框隐藏（如果不可见）
                if not visible and self.selection_rect_item:
                    self.selection_rect_item.setVisible(False)

            def create_table(self):
                """创建可编辑表格"""
                # 清除旧元素
                self.clear_table_elements()
                
                # 重新计算列宽（根据当前节点宽度）
                total_width = self.rect().width() - 2
                self.recalculate_column_widths(total_width)
                
                # 背景 (浅色模式)
                self.setBrush(QBrush(QColor("#ffffff")))
                self.setPen(QPen(QColor("#e0e0e0"), 2))
                
                # 标题区域 (浅色模式)
                title_bg = QGraphicsRectItem(0, 0, self.rect().width(), self.title_height, self)
                title_bg.setBrush(QBrush(QColor("#f8f9fa")))
                title_bg.setPen(QPen(QColor("#e0e0e0"), 1))
                
                # 不可编辑标题（锁定状态）
                self.title_item = QGraphicsTextItem(f"📋 {self.original_table_name}", self)
                self.title_item.setDefaultTextColor(QColor("#202124")) # 深色文字
                self.title_item.setFont(QFont("Microsoft YaHei", 11, QFont.Weight.Bold))
                self.title_item.setPos(15, 10)
                self.title_item.setTextWidth(self.rect().width() - 80)  # 留出按钮空间
                # 标题不可编辑
                self.title_item.setTextInteractionFlags(Qt.TextInteractionFlag.NoTextInteraction)
                
                # 创建折叠/展开按钮（右上角）
                self.toggle_button = ToggleButton(self, self.toggle_collapse)
                self.toggle_button.setPos(self.rect().width() - 40, 7)
                self.toggle_button.update_icon(self.is_collapsed)
                
                # 如果是折叠状态，只显示标题，直接返回
                if self.is_collapsed:
                    return
                
                # 表头区域（展开状态才显示）
                header_y = self.title_height
                header_bg = QGraphicsRectItem(0, header_y, self.rect().width(), self.header_height, self)
                header_bg.setBrush(QBrush(QColor("#f1f3f4"))) # 浅灰背景
                header_bg.setPen(QPen(QColor("#e0e0e0"), 1))
                
                # 创建表头
                x_offset = 1
                self.header_items = []
                for i, (header, width) in enumerate(zip(self.headers, self.column_widths)):
                    # 镜号列不可编辑
                    if i == 0:
                        header_text = QGraphicsTextItem(header, self)
                        header_text.setDefaultTextColor(QColor("#1a73e8")) # Google Blue
                        header_text.setFont(QFont("Microsoft YaHei", 9, QFont.Weight.Bold))
                        header_text.setPos(x_offset + 5, header_y + 10)
                        self.header_items.append(header_text)
                    else:
                        header_item = EditableTextItem(header, self, is_header=True)
                        header_item.setTextWidth(width - 10)
                        header_item.setPos(x_offset + 5, header_y + 10)
                        self.header_items.append(header_item)
                    
                    # 列分隔线
                    if i > 0:
                        line = QGraphicsLineItem(x_offset, header_y, x_offset, header_y + self.header_height, self)
                        line.setPen(QPen(QColor("#dadce0"), 1))
                    
                    x_offset += width
                
                # 创建数据行
                data_y = self.title_height + self.header_height
                self.cell_items = []
                self.row_backgrounds = []
                
                # 创建选中高亮矩形（初始隐藏）
                self.selection_rect_item = QGraphicsRectItem(0, 0, 0, 0, self)
                self.selection_rect_item.setBrush(QBrush(QColor(26, 115, 232, 30))) # 蓝色半透明
                self.selection_rect_item.setPen(QPen(QColor(26, 115, 232), 2)) # 蓝色边框
                self.selection_rect_item.setVisible(False)
                self.selection_rect_item.setZValue(10) # 确保在最上层
                
                # 如果表格为空，显示提示信息
                if len(self.table_data) == 0:
                    empty_hint = QGraphicsTextItem("右键点击表格区域可添加行", self)
                    empty_hint.setDefaultTextColor(QColor("#5f6368"))
                    empty_hint.setFont(QFont("Microsoft YaHei", 9, QFont.Weight.Normal))
                    empty_hint.setPos(self.rect().width() / 2 - 100, data_y + 20)
                    empty_hint.setTextInteractionFlags(Qt.TextInteractionFlag.NoTextInteraction)
                
                for row_idx, row_data in enumerate(self.table_data):
                    y_pos = data_y + (row_idx * self.cell_height)
                    
                    # 行背景（用于高亮选中）
                    row_bg = QGraphicsRectItem(1, y_pos, self.rect().width() - 2, self.cell_height, self)
                    if row_idx % 2 == 0:
                        row_bg.setBrush(QBrush(QColor("#ffffff")))
                    else:
                        row_bg.setBrush(QBrush(QColor("#f8f9fa")))
                    row_bg.setPen(QPen(Qt.PenStyle.NoPen))
                    row_bg.setData(0, row_idx)  # 存储行索引
                    row_bg.setZValue(-1)  # 确保在文字下方
                    self.row_backgrounds.append(row_bg)
                    
                    # 创建单元格
                    x_offset = 1
                    row_cells = []
                    
                    # 第一列：镜头号（不可编辑）
                    shot_number = QGraphicsTextItem(str(row_idx + 1), self)
                    shot_number.setDefaultTextColor(QColor("#5f6368"))
                    shot_number.setFont(QFont("Microsoft YaHei", 9, QFont.Weight.Bold))
                    shot_number.setPos(x_offset + (self.column_widths[0] - 20) / 2, y_pos + 8)
                    row_cells.append(shot_number)
                    x_offset += self.column_widths[0]
                    
                    # 列分隔线
                    line = QGraphicsLineItem(x_offset, y_pos, x_offset, y_pos + self.cell_height, self)
                    line.setPen(QPen(QColor("#f1f3f4"), 1))
                    
                    # 其他列：可编辑单元格（数据列）
                    # ⚠️ 确保即使row_data长度不足，也能显示所有列
                    for col_idx, width in enumerate(self.column_widths[1:]):
                        # 安全获取单元格数据，如果索引超出范围则使用空字符串
                        cell_data = row_data[col_idx] if col_idx < len(row_data) else ""
                        
                        # 判断是否是绘画提示词列（最后两列）
                        # 获取真实的表头索引（col_idx + 1，因为跳过了镜号列）
                        header_idx = col_idx + 1
                        header_name = self.headers[header_idx]
                        
                        is_prompt_column = (header_idx < len(self.headers) and 
                                          ("绘画提示词" in header_name or 
                                           "Drawing Prompt" in header_name))
                        
                        is_draft_column = "草稿" in header_name or "Draft" in header_name
                        is_consistent_image_column = "剧本人物分镜图" in header_name or "一致性分镜图" in header_name or "一致性图片" in header_name
                        # 为草稿列和剧本人物分镜图列使用图片单元格
                        if is_draft_column or is_consistent_image_column:
                            cell_item = ImageCell(
                                str(cell_data), width, self.cell_height, self
                            )
                            cell_item.setPos(x_offset + 2, y_pos + 2)
                            cell_item.setData(0, row_idx)
                            cell_item.setData(1, col_idx)
                            row_cells.append(cell_item)
                        
                        # 为绘画提示词列使用可滚动单元格，其他列使用普通单元格
                        elif is_prompt_column and len(str(cell_data)) > 50:
                            # 可滚动文本单元格（用于长文本）
                            cell_item = ScrollableTextCell(
                                str(cell_data), self, 
                                width, self.cell_height, 
                                is_header=False
                            )
                            cell_item.setPos(x_offset + 5, y_pos + 2)
                            cell_item.setData(0, row_idx)  # 存储行索引
                            cell_item.setData(1, col_idx)  # 存储列索引
                            row_cells.append(cell_item)
                        else:
                            # 普通可编辑文本项（支持长文本省略显示）
                            # 行高增加后，可以显示更多行
                            cell_item = EditableTextItem(str(cell_data), self, is_header=False, max_lines=8)
                            cell_item.setTextWidth(width - 10)
                            cell_item.setPos(x_offset + 5, y_pos + 8)
                            cell_item.setData(0, row_idx)  # 存储行索引
                            cell_item.setData(1, col_idx)  # 存储列索引（不包括镜号）
                            cell_item.update_display()  # 更新显示（应用省略逻辑）
                            row_cells.append(cell_item)
                        
                        x_offset += width
                        
                        # 列分隔线
                        if col_idx < len(self.column_widths) - 2:
                            line = QGraphicsLineItem(x_offset, y_pos, x_offset, y_pos + self.cell_height, self)
                            line.setPen(QPen(QColor("#f1f3f4"), 1))
                    
                    self.cell_items.append(row_cells)
                    
                    # 行分隔线
                    line = QGraphicsLineItem(0, y_pos + self.cell_height, 
                                            self.rect().width(), y_pos + self.cell_height, self)
                    line.setPen(QPen(QColor("#f1f3f4"), 1))

            def refresh_table(self):
                """刷新表格显示"""
                self.create_table()
            
            def recalculate_column_widths(self, total_width):
                """根据总宽度重新计算列宽"""
                # 如果设置了跳过标记，不重新计算列宽
                if hasattr(self, '_skip_column_width_recalc') and self._skip_column_width_recalc:
                    # print(f"[表格节点] 跳过列宽重新计算，保持现有配置: {len(self.column_widths)}列")
                    return
                
                # 保持比例缩放
                # 如果已有列宽配置（包括动态添加的列），则基于当前列宽重新缩放
                if hasattr(self, 'column_widths') and len(self.column_widths) > 0:
                    base_widths = self.column_widths
                    print(f"[表格节点] 使用现有列宽配置: {len(base_widths)}列")
                else:
                    # 默认9列配置
                    base_widths = [80, 100, 250, 120, 150, 120, 100, 250, 120]
                    print(f"[表格节点] 使用默认列宽配置: {len(base_widths)}列")
                
                base_total = sum(base_widths)
                
                if total_width < 500:
                    total_width = 500  # 最小宽度
                
                scale = total_width / base_total
                self.column_widths = [int(w * scale) for w in base_widths]
                # print(f"[表格节点] 重新计算列宽: {len(self.column_widths)}列, 总宽度={total_width}")
            
            def clear_table_elements(self):
                """清除所有表格元素（但保留连接点）"""
                # 导入Socket类用于类型检查
                try:
                    from lingdongconnect import Socket
                    has_socket = True
                except ImportError:
                    has_socket = False
                
                for item in self.childItems():
                    # 保留连接点（Socket），删除其他元素
                    if item != self and hasattr(item, 'scene') and item.scene():
                        # 如果是Socket，跳过不删除
                        if has_socket and isinstance(item, Socket):
                            continue
                        # 删除其他子元素
                        item.scene().removeItem(item)
            
            def mousePressEvent(self, event):
                """鼠标点击事件 - 选中行（标题不可编辑）+ 配置按钮"""
                if event.button() == Qt.MouseButton.LeftButton:
                    pos_y = event.pos().y()
                    pos_x = event.pos().x()
                    
                    # 检查是否点击配置按钮
                    if self.settings_btn_rect and self.settings_btn_rect.contains(event.pos()):
                        self.open_settings_dialog()
                        event.accept()
                        return
                    
                    # 标题区域不响应点击（用于拖动节点）
                    if pos_y < self.title_height:
                        # 继续默认行为，允许拖动节点
                        pass
                    
                    # 检查是否点击在缩放手柄上
                    if self.is_in_resize_handle(event.pos()):
                        self.is_resizing = True
                        self.resize_start_pos = event.pos()
                        event.accept()
                        return
                    
                    # 检查点击的行（表头下方的数据区）
                    data_start_y = self.title_height + self.header_height
                    
                    if pos_y >= data_start_y:
                        row_idx = int((pos_y - data_start_y) / self.cell_height)
                        if 0 <= row_idx < len(self.table_data):
                            # 计算列索引
                            x_offset = 1
                            col_idx = -1
                            for i, width in enumerate(self.column_widths):
                                if pos_x >= x_offset and pos_x < x_offset + width:
                                    col_idx = i
                                    break
                                x_offset += width

                            if col_idx >= 0:
                                self.select_cell(row_idx, col_idx)
                        
                        # 确保点击表格内容也能选中节点
                        self.setSelected(True)
                        
                        event.accept()
                        return
                
                super().mousePressEvent(event)
            
            def mouseMoveEvent(self, event):
                """鼠标移动事件 - 缩放"""
                if self.is_resizing and self.resize_start_pos:
                    delta = event.pos() - self.resize_start_pos
                    
                    # 计算新宽度（保持最小宽度）
                    new_width = max(500, self.rect().width() + delta.x())
                    
                    # 高度根据行数自动计算，不允许手动调整
                    row_count = len(self.table_data)
                    new_height = self.title_height + self.header_height + (row_count * self.cell_height) + 2
                    
                    # 更新大小
                    self.setRect(0, 0, new_width, new_height)
                    
                    # 重新创建表格
                    self.create_table()
                    
                    self.resize_start_pos = event.pos()
                    event.accept()
                else:
                    super().mouseMoveEvent(event)
            
            def mouseReleaseEvent(self, event):
                """鼠标释放事件"""
                if self.is_resizing:
                    self.is_resizing = False
                    self.resize_start_pos = None
                    event.accept()
                else:
                    super().mouseReleaseEvent(event)
            
            def is_in_resize_handle(self, pos):
                """检查是否在缩放手柄区域"""
                rect = self.rect()
                handle_rect = QRectF(
                    rect.width() - self.resize_handle_size,
                    rect.height() - self.resize_handle_size,
                    self.resize_handle_size,
                    self.resize_handle_size
                )
                return handle_rect.contains(pos)
            
            def select_cell(self, row_idx, col_idx=-1):
                """选中单元格或行（红色半透明高亮）"""
                self.selected_row = row_idx
                self.selected_col = col_idx
                
                if row_idx >= 0 and row_idx < len(self.table_data):
                    # 计算高亮区域位置和大小
                    y_pos = self.title_height + self.header_height + (row_idx * self.cell_height)
                    
                    if col_idx >= 0:
                        # 选中特定单元格
                        x_offset = 1
                        for i in range(col_idx):
                            x_offset += self.column_widths[i]
                        
                        width = self.column_widths[col_idx]
                        
                        # 更新高亮矩形
                        if self.selection_rect_item:
                            self.selection_rect_item.setRect(x_offset, y_pos, width, self.cell_height)
                            self.selection_rect_item.setVisible(True)
                            print(f"[表格节点] 选中单元格: 行{row_idx + 1}, 列{col_idx + 1}")
                    else:
                        # 选中整行
                        if self.selection_rect_item:
                            self.selection_rect_item.setRect(1, y_pos, self.rect().width() - 2, self.cell_height)
                            self.selection_rect_item.setVisible(True)
                            print(f"[表格节点] 选中第 {row_idx + 1} 行")
                else:
                    # 隐藏高亮
                    if self.selection_rect_item:
                        self.selection_rect_item.setVisible(False)

            def select_row(self, row_idx):
                """兼容旧接口"""
                self.select_cell(row_idx, -1)
            
            def contextMenuEvent(self, event):
                """右键菜单 - 支持复制和修改单元格 + 标题栏下载/上传"""
                pos_y = event.pos().y()
                pos_x = event.pos().x()
                
                menu = QMenu()
                menu.setStyleSheet("""
                    QMenu {
                        background-color: #ffffff;
                        color: #202124;
                        border: 1px solid #dadce0;
                        padding: 5px;
                    }
                    QMenu::item {
                        padding: 8px 20px;
                    }
                    QMenu::item:selected {
                        background-color: #e8f0fe;
                        color: #1967d2;
                    }
                """)
                
                # 如果右键点击标题区域，显示下载和上传选项
                if pos_y < self.title_height:
                    download_action = QAction("📥 下载为TXT", None)
                    download_action.triggered.connect(self.download_txt)
                    menu.addAction(download_action)
                    
                    upload_action = QAction("📤 上传TXT", None)
                    upload_action.triggered.connect(self.upload_txt)
                    menu.addAction(upload_action)
                else:
                    # 检查是否点击在数据单元格上
                    data_start_y = self.title_height + self.header_height
                    if pos_y >= data_start_y and not self.is_collapsed:
                        # 计算点击的行和列
                        row_idx = int((pos_y - data_start_y) / self.cell_height)
                        
                        if 0 <= row_idx < len(self.table_data):
                            # 计算列索引
                            x_offset = 1
                            col_idx = -1
                            for i, width in enumerate(self.column_widths):
                                if pos_x >= x_offset and pos_x < x_offset + width:
                                    col_idx = i
                                    break
                                x_offset += width
                            
                            # 如果点击在数据列上（排除镜号列）
                            if col_idx > 0:
                                data_col_idx = col_idx - 1  # 数据列索引
                                
                                # 获取单元格内容
                                if data_col_idx < len(self.table_data[row_idx]):
                                    cell_content = str(self.table_data[row_idx][data_col_idx])
                                    
                                    # 添加复制菜单
                                    copy_action = QAction(f"📋 复制", None)
                                    copy_action.triggered.connect(lambda: self.copy_cell_content(row_idx, data_col_idx))
                                    menu.addAction(copy_action)
                                    
                                    # 添加修改菜单
                                    edit_action = QAction(f"✏️ 修改", None)
                                    edit_action.triggered.connect(lambda: self.edit_cell_content(row_idx, data_col_idx))
                                    menu.addAction(edit_action)
                                    
                                    # 添加分隔符
                                    menu.addSeparator()
                                    
                                    # 检查是否是"剧本人物分镜图"列
                                    is_consistent_col = False
                                    if col_idx < len(self.headers):
                                        header_name = self.headers[col_idx]
                                        if "剧本人物分镜图" in header_name or "一致性分镜图" in header_name or "Consistent Storyboard" in header_name:
                                            is_consistent_col = True
                                    
                                    # 添加重新生成图片菜单（仅在剧本人物分镜图列显示）
                                    if is_consistent_col:
                                        regenerate_action = QAction(f"🎨 重新生成第{row_idx+1}镜图片", None)
                                        regenerate_action.triggered.connect(lambda: self.regenerate_single_image(row_idx))
                                        menu.addAction(regenerate_action)
                
                # 只有在菜单有项目时才显示
                if not menu.isEmpty():
                    menu.exec(event.screenPos())
                
                event.accept()
            
            def copy_cell_content(self, row_idx, col_idx):
                """复制单元格内容到剪贴板"""
                try:
                    from PySide6.QtWidgets import QApplication
                    
                    if 0 <= row_idx < len(self.table_data) and 0 <= col_idx < len(self.table_data[row_idx]):
                        cell_content = str(self.table_data[row_idx][col_idx])
                        
                        # 复制到剪贴板
                        clipboard = QApplication.clipboard()
                        clipboard.setText(cell_content)
                        
                        print(f"[表格节点] 已复制单元格内容: 第{row_idx+1}行, 第{col_idx+2}列")
                        print(f"[表格节点] 内容: {cell_content[:100]}{'...' if len(cell_content) > 100 else ''}")
                        
                except Exception as e:
                    print(f"[表格节点] 复制失败: {e}")
                    import traceback
                    traceback.print_exc()
            
            def open_settings_dialog(self):
                """打开配置对话框"""
                try:
                    from lingdongsetting import StoryboardSettingsDialog
                    
                    # 当前配置
                    current_settings = {
                        'cell_height': self.cell_height,
                        'header_height': self.header_height,
                        'auto_number': True,
                        'highlight_color': '#ff0000',
                        'text_color': '#cccccc',
                        'header_color': '#00bfff',
                        'show_grid': True,
                        'max_text_length': 80
                    }
                    
                    # 创建并显示对话框
                    dialog = StoryboardSettingsDialog(None, current_settings)
                    
                    if dialog.exec():
                        # 应用新配置
                        new_settings = dialog.get_settings()
                        self.cell_height = new_settings['cell_height']
                        self.header_height = new_settings['header_height']
                        
                        # 刷新表格
                        self.refresh_table()
                        
                        print(f"[表格配置] 已更新: 单元格高度={self.cell_height}, 表头高度={self.header_height}")
                
                except Exception as e:
                    print(f"[表格配置] 打开配置对话框失败: {e}")
                    import traceback
                    traceback.print_exc()
            
            def download_txt(self):
                """下载表格为TXT文本（表格形式）"""
                try:
                    # 收集当前数据
                    self.collect_data()
                    
                    # 构建TXT内容（表格形式，使用制表符分隔）
                    txt_content = ""
                    
                    # 表头行
                    txt_content += "\t".join(self.headers) + "\n"
                    
                    # 数据行
                    for idx, row in enumerate(self.table_data):
                        # 镜号
                        row_data = [str(idx + 1)]
                        # 其他列
                        row_data.extend([str(cell) for cell in row])
                        txt_content += "\t".join(row_data) + "\n"
                    
                    # 保存文件
                    from PySide6.QtWidgets import QFileDialog
                    file_path, _ = QFileDialog.getSaveFileName(
                        None,
                        "保存TXT文件",
                        f"{self.original_table_name}.txt",
                        "文本文件 (*.txt);;所有文件 (*.*)"
                    )
                    
                    if file_path:
                        with open(file_path, 'w', encoding='utf-8') as f:
                            f.write(txt_content)
                        
                        print(f"[表格节点] 已下载为TXT: {file_path}")
                        print(f"[表格节点] 共 {len(self.table_data)} 行数据")
                
                except Exception as e:
                    print(f"[表格节点] 下载TXT失败: {e}")
                    import traceback
                    traceback.print_exc()
            
            def upload_txt(self):
                """上传TXT文本并导入到表格"""
                try:
                    from PySide6.QtWidgets import QFileDialog, QMessageBox
                    
                    # 选择文件
                    file_path, _ = QFileDialog.getOpenFileName(
                        None,
                        "选择TXT文件",
                        "",
                        "文本文件 (*.txt);;所有文件 (*.*)"
                    )
                    
                    if not file_path:
                        return
                    
                    # 读取文件
                    with open(file_path, 'r', encoding='utf-8') as f:
                        lines = f.readlines()
                    
                    if len(lines) < 2:
                        QMessageBox.warning(None, "错误", "TXT文件格式不正确，至少需要表头和一行数据")
                        return
                    
                    # 解析表头
                    header_line = lines[0].strip()
                    new_headers = [h.strip() for h in header_line.split('\t')]
                    
                    # 解析数据
                    new_data = []
                    for line in lines[1:]:
                        line = line.strip()
                        if not line:
                            continue
                        
                        cells = line.split('\t')
                        # 跳过第一列（镜号，会自动生成）
                        if len(cells) > 1:
                            new_data.append(cells[1:])
                    
                    # 确认导入
                    reply = QMessageBox.question(
                        None,
                        "确认导入",
                        f"将导入 {len(new_data)} 行数据，{len(new_headers)} 列\n\n是否继续？",
                        QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
                    )
                    
                    if reply == QMessageBox.StandardButton.Yes:
                        # 更新表头和数据
                        self.headers = new_headers
                        self.table_data = new_data
                        
                        # 重新计算列宽
                        num_cols = len(new_headers)
                        if num_cols <= 5:
                            self.column_widths = [60] + [150] * (num_cols - 1)
                        else:
                            self.column_widths = [60] + [120] * (num_cols - 1)
                        
                        # 刷新表格
                        self.refresh_table()
                        
                        print(f"[表格节点] 已导入TXT: {file_path}")
                        print(f"[表格节点] 表头: {new_headers}")
                        print(f"[表格节点] 数据行数: {len(new_data)}")
                
                except Exception as e:
                    print(f"[表格节点] 上传TXT失败: {e}")
                    import traceback
                    traceback.print_exc()
                    QMessageBox.critical(None, "错误", f"导入失败: {str(e)}")
            
            def edit_cell_content(self, row_idx, col_idx):
                """修改单元格内容"""
                try:
                    if 0 <= row_idx < len(self.table_data) and 0 <= col_idx < len(self.table_data[row_idx]):
                        current_content = str(self.table_data[row_idx][col_idx])
                        
                        # 获取列标题
                        header_idx = col_idx + 1  # 加1是因为跳过了镜号列
                        column_name = self.headers[header_idx] if header_idx < len(self.headers) else f"第{col_idx+2}列"
                        
                        # 使用自定义编辑对话框
                        dialog = CellEditDialog(
                            f"修改单元格 - 第{row_idx+1}行",
                            f"{column_name}:",
                            current_content
                        )
                        
                        if dialog.exec():
                            text = dialog.text
                            # 更新数据
                            self.table_data[row_idx][col_idx] = text
                            
                            # 刷新表格显示
                            self.refresh_table()
                            
                            print(f"[表格节点] 已修改单元格: 第{row_idx+1}行, {column_name}")
                            print(f"[表格节点] 新内容: {text[:100]}{'...' if len(text) > 100 else ''}")
                
                except Exception as e:
                    print(f"[表格节点] 修改失败: {e}")
                    import traceback
                    traceback.print_exc()
            
            def regenerate_single_image(self, row_idx):
                """重新生成单个镜头的图片"""
                try:
                    print(f"\n[单镜重生成] ========== 开始重新生成第{row_idx+1}镜 ==========")
                    
                    # 获取主界面引用
                    if not hasattr(self, 'scene') or not self.scene():
                        print(f"[单镜重生成] ✗ 无法获取场景引用")
                        return
                    
                    canvas = self.scene()
                    if not hasattr(canvas, 'main_page'):
                        print(f"[单镜重生成] ✗ 无法获取主界面引用")
                        return
                    
                    main_page = canvas.main_page
                    
                    # 调用主界面的单镜生成方法
                    # if hasattr(main_page, 'workbench') and hasattr(main_page.workbench, 'regenerate_single_shot'):
                    #     main_page.workbench.regenerate_single_shot(self, row_idx)
                    # elif hasattr(main_page, 'regenerate_single_shot'):
                    #     main_page.regenerate_single_shot(self, row_idx)
                    # else:
                    #     print(f"[单镜重生成] ✗ 主界面及其工作台没有regenerate_single_shot方法")
                    print(f"[单镜重生成] 功能已移除")
                
                except Exception as e:
                    print(f"[单镜重生成] ✗ 失败: {e}")
                    import traceback
                    traceback.print_exc()
            
            def open_settings_dialog(self):
                """打开配置对话框"""
                try:
                    from lingdongsetting import StoryboardSettingsDialog
                    
                    # 当前配置
                    current_settings = {
                        'cell_height': self.cell_height,
                        'header_height': self.header_height,
                        'auto_number': True,
                        'highlight_color': '#ff0000',
                        'text_color': '#cccccc',
                        'header_color': '#00bfff',
                        'show_grid': True,
                        'max_text_length': 80
                    }
                    
                    # 创建并显示对话框
                    dialog = StoryboardSettingsDialog(None, current_settings)
                    
                    if dialog.exec():
                        # 应用新配置
                        new_settings = dialog.get_settings()
                        self.cell_height = new_settings['cell_height']
                        self.header_height = new_settings['header_height']
                        
                        # 刷新表格
                        self.refresh_table()
                        
                        print(f"[表格配置] 已更新: 单元格高度={self.cell_height}, 表头高度={self.header_height}")
                
                except Exception as e:
                    print(f"[表格配置] 打开配置对话框失败: {e}")
                    import traceback
                    traceback.print_exc()
            
            def download_txt(self):
                """下载表格为TXT文本（表格形式）"""
                try:
                    # 收集当前数据
                    self.collect_data()
                    
                    # 构建TXT内容（表格形式，使用制表符分隔）
                    txt_content = ""
                    
                    # 表头行
                    txt_content += "\t".join(self.headers) + "\n"
                    
                    # 数据行
                    for idx, row in enumerate(self.table_data):
                        # 镜号
                        row_data = [str(idx + 1)]
                        # 其他列
                        row_data.extend([str(cell) for cell in row])
                        txt_content += "\t".join(row_data) + "\n"
                    
                    # 保存文件
                    from PySide6.QtWidgets import QFileDialog
                    file_path, _ = QFileDialog.getSaveFileName(
                        None,
                        "保存TXT文件",
                        f"{self.original_table_name}.txt",
                        "文本文件 (*.txt);;所有文件 (*.*)"
                    )
                    
                    if file_path:
                        with open(file_path, 'w', encoding='utf-8') as f:
                            f.write(txt_content)
                        
                        print(f"[表格节点] 已下载为TXT: {file_path}")
                        print(f"[表格节点] 共 {len(self.table_data)} 行数据")
                
                except Exception as e:
                    print(f"[表格节点] 下载TXT失败: {e}")
                    import traceback
                    traceback.print_exc()
            
            def upload_txt(self):
                """上传TXT文本并导入到表格"""
                try:
                    from PySide6.QtWidgets import QFileDialog, QMessageBox
                    
                    # 选择文件
                    file_path, _ = QFileDialog.getOpenFileName(
                        None,
                        "选择TXT文件",
                        "",
                        "文本文件 (*.txt);;所有文件 (*.*)"
                    )
                    
                    if not file_path:
                        return
                    
                    # 读取文件
                    with open(file_path, 'r', encoding='utf-8') as f:
                        lines = f.readlines()
                    
                    if len(lines) < 2:
                        QMessageBox.warning(None, "错误", "TXT文件格式不正确，至少需要表头和一行数据")
                        return
                    
                    # 解析表头
                    header_line = lines[0].strip()
                    new_headers = [h.strip() for h in header_line.split('\t')]
                    
                    # 解析数据
                    new_data = []
                    for line in lines[1:]:
                        line = line.strip()
                        if not line:
                            continue
                        
                        cells = line.split('\t')
                        # 跳过第一列（镜号，会自动生成）
                        if len(cells) > 1:
                            new_data.append(cells[1:])
                    
                    # 确认导入
                    reply = QMessageBox.question(
                        None,
                        "确认导入",
                        f"将导入 {len(new_data)} 行数据，{len(new_headers)} 列\n\n是否继续？",
                        QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
                    )
                    
                    if reply == QMessageBox.StandardButton.Yes:
                        # 更新表头和数据
                        self.headers = new_headers
                        self.table_data = new_data
                        
                        # 重新计算列宽
                        num_cols = len(new_headers)
                        if num_cols <= 5:
                            self.column_widths = [60] + [150] * (num_cols - 1)
                        else:
                            self.column_widths = [60] + [120] * (num_cols - 1)
                        
                        # 刷新表格
                        self.refresh_table()
                        
                        print(f"[表格节点] 已导入TXT: {file_path}")
                        print(f"[表格节点] 表头: {new_headers}")
                        print(f"[表格节点] 数据行数: {len(new_data)}")
                
                except Exception as e:
                    print(f"[表格节点] 上传TXT失败: {e}")
                    import traceback
                    traceback.print_exc()
                    QMessageBox.critical(None, "错误", f"导入失败: {str(e)}")
            
            def add_row(self):
                """添加新行"""
                self.table_data.append(["", "", "", "", "", "", "", ""])
                self.refresh_table()
            
            def delete_row(self, row_idx):
                """删除指定行"""
                if 0 <= row_idx < len(self.table_data):
                    self.table_data.pop(row_idx)
                    self.selected_row = -1
                    self.refresh_table()
            
            def refresh_table(self):
                """刷新表格显示"""
                # 收集当前数据
                self.collect_data()
                
                # 重新计算大小
                total_width = sum(self.column_widths) + 2
                row_count = len(self.table_data)
                total_height = self.title_height + self.header_height + (row_count * self.cell_height) + 2
                
                # 保存展开状态的高度
                self.expanded_height = total_height
                
                # 如果是折叠状态，使用折叠高度
                if self.is_collapsed:
                    self.setRect(0, 0, total_width, self.collapsed_height)
                else:
                    self.setRect(0, 0, total_width, total_height)
                
                # 重新创建表格
                self.create_table()
                
                # ⭐ 更新连接点位置（如果节点支持连接功能）
                if hasattr(self, 'update_socket_positions'):
                    self.update_socket_positions()
                
                # ⭐ 更新所有连接线（如果节点支持连接功能）
                if hasattr(self, 'update_connections'):
                    self.update_connections()
            
            def toggle_collapse(self):
                """切换折叠/展开状态"""
                self.is_collapsed = not self.is_collapsed
                
                # 收集当前数据
                self.collect_data()
                
                # 计算尺寸
                total_width = sum(self.column_widths) + 2
                
                if self.is_collapsed:
                    # 折叠：只显示标题栏
                    self.setRect(0, 0, total_width, self.collapsed_height)
                    print(f"[表格节点] 折叠表格: {self.original_table_name}")
                else:
                    # 展开：显示完整表格
                    row_count = len(self.table_data)
                    total_height = self.title_height + self.header_height + (row_count * self.cell_height) + 2
                    self.expanded_height = total_height
                    self.setRect(0, 0, total_width, total_height)
                    print(f"[表格节点] 展开表格: {self.original_table_name}")
                
                # 重新创建表格
                self.create_table()
                
                # ⭐ 更新连接点位置（如果节点支持连接功能）
                if hasattr(self, 'update_socket_positions'):
                    self.update_socket_positions()
                    print(f"[表格节点] 已更新连接点位置")
                
                # ⭐ 更新所有连接线（如果节点支持连接功能）
                if hasattr(self, 'update_connections'):
                    self.update_connections()
                    print(f"[表格节点] 已更新连接线")
            
            def collect_data(self):
                """收集当前编辑的数据"""
                # 标题不可编辑，无需收集
                # self.original_table_name 由系统自动管理
                
                # 收集表头（跳过镜号列）
                for i in range(1, len(self.header_items)):
                    if hasattr(self.header_items[i], 'toPlainText'):
                        if i < len(self.headers):
                            self.headers[i] = self.header_items[i].toPlainText().strip()
                
                # 收集单元格数据（跳过镜号列，从第二列开始）
                for row_idx, row_cells in enumerate(self.cell_items):
                    if row_idx < len(self.table_data):
                        # row_cells[0] 是镜号，跳过
                        # row_cells[1:] 是实际数据（8列）
                        for col_idx in range(1, len(row_cells)):
                            data_col_idx = col_idx - 1  # 数据列索引（0-7）
                            if data_col_idx < len(self.table_data[row_idx]):
                                if hasattr(row_cells[col_idx], 'toPlainText'):
                                    self.table_data[row_idx][data_col_idx] = row_cells[col_idx].toPlainText().strip()
            
            def paint(self, painter, option, widget=None):
                """绘制节点（添加缩放手柄和配置按钮）"""
                super().paint(painter, option, widget)
                
                # 绘制配置按钮（标题栏右侧，折叠按钮左边）
                rect = self.rect()
                settings_x = rect.width() - 60  # 折叠按钮在width-30，配置按钮在其左边30
                settings_y = 12
                
                self.settings_btn_rect = QRectF(
                    settings_x, settings_y,
                    self.settings_btn_size, self.settings_btn_size
                )
                
                # 绘制配置图标（齿轮图标）
                painter.setRenderHint(QPainter.RenderHint.Antialiasing)
                painter.setPen(QPen(QColor("#888888"), 1.5))
                painter.setBrush(Qt.BrushStyle.NoBrush)
                
                # 简化的齿轮图标（一个圆圈加中心点）
                center_x = settings_x + self.settings_btn_size / 2
                center_y = settings_y + self.settings_btn_size / 2
                painter.drawEllipse(QRectF(
                    center_x - 7, center_y - 7, 14, 14
                ))
                painter.drawEllipse(QRectF(
                    center_x - 3, center_y - 3, 6, 6
                ))
                
                # 绘制缩放手柄（右下角）
                if self.isSelected():
                    painter.setBrush(QBrush(QColor("#00bfff")))
                    painter.setPen(QPen(QColor("#00bfff"), 1))
                    handle_rect = QRectF(
                        rect.width() - self.resize_handle_size,
                        rect.height() - self.resize_handle_size,
                        self.resize_handle_size,
                        self.resize_handle_size
                    )
                    painter.drawRect(handle_rect)
            
            def set_table_data(self, data, name="分镜脚本"):
                """设置表格数据（用于从AI生成的表格创建）"""
                self.table_data = data
                self.original_table_name = name
                self.node_title = name
                self.refresh_table()
            
            def open_in_browser(self):
                """在浏览器中打开表格"""
                import webbrowser
                import tempfile
                from datetime import datetime
                
                try:
                    # 收集当前数据
                    self.collect_data()
                    
                    # 生成HTML内容
                    html_content = f"""<!DOCTYPE html>
<html lang="zh-CN">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>{self.original_table_name}</title>
    <style>
        * {{
            margin: 0;
            padding: 0;
            box-sizing: border-box;
        }}
        
        body {{
            font-family: 'Microsoft YaHei', 'PingFang SC', -apple-system, BlinkMacSystemFont, sans-serif;
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            min-height: 100vh;
            padding: 40px 20px;
        }}
        
        .container {{
            max-width: 1400px;
            margin: 0 auto;
            background: white;
            border-radius: 16px;
            overflow: hidden;
        }}
        
        .header {{
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            color: white;
            padding: 30px 40px;
            display: flex;
            justify-content: space-between;
            align-items: center;
        }}
        
        .header h1 {{
            font-size: 28px;
            font-weight: 600;
            display: flex;
            align-items: center;
            gap: 12px;
        }}
        
        .header .info {{
            font-size: 14px;
            opacity: 0.9;
        }}
        
        .table-wrapper {{
            padding: 40px;
            overflow-x: auto;
        }}
        
        table {{
            width: 100%;
            border-collapse: separate;
            border-spacing: 0;
            border-radius: 8px;
            overflow: hidden;
        }}
        
        th {{
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            color: white;
            font-weight: 600;
            text-align: center;
            padding: 16px 12px;
            font-size: 14px;
            white-space: nowrap;
        }}
        
        th:first-child {{
            border-top-left-radius: 8px;
        }}
        
        th:last-child {{
            border-top-right-radius: 8px;
        }}
        
        td {{
            padding: 14px 12px;
            border-bottom: 1px solid #e8e8e8;
            font-size: 13px;
            color: #333;
            vertical-align: top;
        }}
        
        tr:last-child td {{
            border-bottom: none;
        }}
        
        tr:nth-child(even) {{
            background-color: #f9f9f9;
        }}
        
        tr:hover {{
            background-color: #f0f4ff;
            transition: background-color 0.2s ease;
        }}
        
        td:first-child {{
            text-align: center;
            font-weight: 600;
            color: #667eea;
            background-color: #f5f7ff;
        }}
        
        .empty-state {{
            text-align: center;
            padding: 60px 20px;
            color: #999;
        }}
        
        .footer {{
            text-align: center;
            padding: 20px;
            color: #999;
            font-size: 12px;
            border-top: 1px solid #e8e8e8;
        }}
        
        @media print {{
            body {{
                background: white;
                padding: 0;
            }}
            
            .container {{
            }}
            
            .header {{
                background: #667eea !important;
                -webkit-print-color-adjust: exact;
                print-color-adjust: exact;
            }}
            
            th {{
                background: #667eea !important;
                -webkit-print-color-adjust: exact;
                print-color-adjust: exact;
            }}
            
            .footer {{
                display: none;
            }}
        }}
    </style>
</head>
<body>
    <div class="container">
        <div class="header">
            <h1>
                <span>📋</span>
                {self.original_table_name}
            </h1>
            <div class="info">
                生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}<br>
                共 {len(self.table_data)} 条记录
            </div>
        </div>
        
        <div class="table-wrapper">
"""
                    
                    # 添加表格
                    if self.table_data:
                        html_content += "            <table>\n"
                        html_content += "                <thead>\n                    <tr>\n"
                        
                        # 表头
                        for header in self.headers:
                            html_content += f"                        <th>{header}</th>\n"
                        
                        html_content += "                    </tr>\n                </thead>\n"
                        html_content += "                <tbody>\n"
                        
                        # 数据行
                        for row_idx, row_data in enumerate(self.table_data, start=1):
                            html_content += "                    <tr>\n"
                            
                            # 第一列：镜号
                            html_content += f"                        <td>{row_idx}</td>\n"
                            
                            # 其他列：数据
                            for cell_value in row_data:
                                # HTML转义
                                cell_text = str(cell_value).replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;')
                                html_content += f"                        <td>{cell_text}</td>\n"
                            
                            html_content += "                    </tr>\n"
                        
                        html_content += "                </tbody>\n            </table>\n"
                    else:
                        html_content += """            <div class="empty-state">
                <p style="font-size: 48px; margin-bottom: 16px;">📭</p>
                <p style="font-size: 16px;">暂无数据</p>
            </div>
"""
                    
                    html_content += """        </div>
        
        <div class="footer">
            由灵动智能体生成 | 按 Ctrl+P 可打印此页面
        </div>
    </div>
</body>
</html>"""
                    
                    # 创建临时HTML文件
                    with tempfile.NamedTemporaryFile(mode='w', encoding='utf-8', suffix='.html', delete=False) as f:
                        f.write(html_content)
                        temp_file = f.name
                    
                    # 在浏览器中打开
                    webbrowser.open('file://' + temp_file)
                    
                    print(f"[表格节点] 已在浏览器中打开: {temp_file}")
                    
                except Exception as e:
                    print(f"[表格节点] 打开网页失败: {e}")
                    import traceback
                    traceback.print_exc()
            
            def open_settings_dialog(self):
                """打开配置对话框"""
                try:
                    from lingdongsetting import StoryboardSettingsDialog
                    
                    # 当前配置
                    current_settings = {
                        'cell_height': self.cell_height,
                        'header_height': self.header_height,
                        'auto_number': True,
                        'highlight_color': '#ff0000',
                        'text_color': '#cccccc',
                        'header_color': '#00bfff',
                        'show_grid': True,
                        'max_text_length': 80
                    }
                    
                    # 创建并显示对话框
                    dialog = StoryboardSettingsDialog(None, current_settings)
                    
                    if dialog.exec():
                        # 应用新配置
                        new_settings = dialog.get_settings()
                        self.cell_height = new_settings['cell_height']
                        self.header_height = new_settings['header_height']
                        
                        # 刷新表格
                        self.refresh_table()
                        
                        print(f"[表格配置] 已更新: 单元格高度={self.cell_height}, 表头高度={self.header_height}")
                
                except Exception as e:
                    print(f"[表格配置] 打开配置对话框失败: {e}")
                    import traceback
                    traceback.print_exc()
            
            def download_txt(self):
                """下载表格为TXT文本（表格形式）"""
                try:
                    # 收集当前数据
                    self.collect_data()
                    
                    # 构建TXT内容（表格形式，使用制表符分隔）
                    txt_content = ""
                    
                    # 表头行
                    txt_content += "\t".join(self.headers) + "\n"
                    
                    # 数据行
                    for idx, row in enumerate(self.table_data):
                        # 镜号
                        row_data = [str(idx + 1)]
                        # 其他列
                        row_data.extend([str(cell) for cell in row])
                        txt_content += "\t".join(row_data) + "\n"
                    
                    # 保存文件
                    from PySide6.QtWidgets import QFileDialog
                    file_path, _ = QFileDialog.getSaveFileName(
                        None,
                        "保存TXT文件",
                        f"{self.original_table_name}.txt",
                        "文本文件 (*.txt);;所有文件 (*.*)"
                    )
                    
                    if file_path:
                        with open(file_path, 'w', encoding='utf-8') as f:
                            f.write(txt_content)
                        
                        print(f"[表格节点] 已下载为TXT: {file_path}")
                        print(f"[表格节点] 共 {len(self.table_data)} 行数据")
                
                except Exception as e:
                    print(f"[表格节点] 下载TXT失败: {e}")
                    import traceback
                    traceback.print_exc()
            
            def upload_txt(self):
                """上传TXT文本并导入到表格"""
                try:
                    from PySide6.QtWidgets import QFileDialog, QMessageBox
                    
                    # 选择文件
                    file_path, _ = QFileDialog.getOpenFileName(
                        None,
                        "选择TXT文件",
                        "",
                        "文本文件 (*.txt);;所有文件 (*.*)"
                    )
                    
                    if not file_path:
                        return
                    
                    # 读取文件
                    with open(file_path, 'r', encoding='utf-8') as f:
                        lines = f.readlines()
                    
                    if len(lines) < 2:
                        QMessageBox.warning(None, "错误", "TXT文件格式不正确，至少需要表头和一行数据")
                        return
                    
                    # 解析表头
                    header_line = lines[0].strip()
                    new_headers = [h.strip() for h in header_line.split('\t')]
                    
                    # 解析数据
                    new_data = []
                    for line in lines[1:]:
                        line = line.strip()
                        if not line:
                            continue
                        
                        cells = line.split('\t')
                        # 跳过第一列（镜号，会自动生成）
                        if len(cells) > 1:
                            new_data.append(cells[1:])
                    
                    # 确认导入
                    reply = QMessageBox.question(
                        None,
                        "确认导入",
                        f"将导入 {len(new_data)} 行数据，{len(new_headers)} 列\n\n是否继续？",
                        QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
                    )
                    
                    if reply == QMessageBox.StandardButton.Yes:
                        # 更新表头和数据
                        self.headers = new_headers
                        self.table_data = new_data
                        
                        # 重新计算列宽
                        num_cols = len(new_headers)
                        if num_cols <= 5:
                            self.column_widths = [60] + [150] * (num_cols - 1)
                        else:
                            self.column_widths = [60] + [120] * (num_cols - 1)
                        
                        # 刷新表格
                        self.refresh_table()
                        
                        print(f"[表格节点] 已导入TXT: {file_path}")
                        print(f"[表格节点] 表头: {new_headers}")
                        print(f"[表格节点] 数据行数: {len(new_data)}")
                
                except Exception as e:
                    print(f"[表格节点] 上传TXT失败: {e}")
                    import traceback
                    traceback.print_exc()
                    QMessageBox.critical(None, "错误", f"导入失败: {str(e)}")
                    QMessageBox.warning(None, "错误", f"无法在浏览器中打开:\n{str(e)}")
            
            def upload_text_file(self):
                """上传文本文件到分镜脚本节点"""
                try:
                    # 弹出文件选择对话框
                    file_path, _ = QFileDialog.getOpenFileName(
                        None,
                        "选择文本文件",
                        "",
                        "文本文件 (*.txt);;所有文件 (*.*)"
                    )
                    
                    if not file_path:
                        return
                    
                    # 读取文件内容
                    with open(file_path, 'r', encoding='utf-8') as f:
                        content = f.read()
                    
                    if not content.strip():
                        QMessageBox.warning(None, "警告", "文本文件为空！")
                        return
                    
                    # 解析文本内容为表格数据
                    # 支持两种格式：
                    # 1. 每行一条数据，填充到"画面内容"列
                    # 2. 制表符分隔的多列数据
                    lines = [line.strip() for line in content.strip().split('\n') if line.strip()]
                    
                    if not lines:
                        QMessageBox.warning(None, "警告", "没有有效的文本行！")
                        return
                    
                    # 清空现有数据
                    self.table_data = []
                    
                    # 解析每一行
                    for line in lines:
                        # 检查是否包含制表符（多列格式）
                        if '\t' in line:
                            # 制表符分隔的多列数据
                            parts = line.split('\t')
                            # 确保有8列数据（去掉镜号列）
                            row_data = []
                            for i in range(8):
                                if i < len(parts):
                                    row_data.append(parts[i])
                                else:
                                    row_data.append("")
                            self.table_data.append(row_data)
                        else:
                            # 单列格式：全部填充到"画面内容"列（第3列，索引1）
                            row_data = ["", line, "", "", "", "", "", ""]
                            self.table_data.append(row_data)
                    
                    # 刷新表格显示
                    self.refresh_table()
                    
                    print(f"[表格节点] 成功导入 {len(self.table_data)} 行数据")
                    QMessageBox.information(
                        None, 
                        "导入成功", 
                        f"成功导入 {len(self.table_data)} 行数据！\n\n提示：\n- 单行文本将填充到'画面内容'列\n- 使用制表符(Tab)分隔可导入多列数据"
                    )
                    
                except Exception as e:
                    print(f"[表格节点] 上传文本失败: {e}")
                    import traceback
                    traceback.print_exc()
            
            def open_settings_dialog(self):
                """打开配置对话框"""
                try:
                    from lingdongsetting import StoryboardSettingsDialog
                    
                    # 当前配置
                    current_settings = {
                        'cell_height': self.cell_height,
                        'header_height': self.header_height,
                        'auto_number': True,
                        'highlight_color': '#ff0000',
                        'text_color': '#cccccc',
                        'header_color': '#00bfff',
                        'show_grid': True,
                        'max_text_length': 80
                    }
                    
                    # 创建并显示对话框
                    dialog = StoryboardSettingsDialog(None, current_settings)
                    
                    if dialog.exec():
                        # 应用新配置
                        new_settings = dialog.get_settings()
                        self.cell_height = new_settings['cell_height']
                        self.header_height = new_settings['header_height']
                        
                        # 刷新表格
                        self.refresh_table()
                        
                        print(f"[表格配置] 已更新: 单元格高度={self.cell_height}, 表头高度={self.header_height}")
                
                except Exception as e:
                    print(f"[表格配置] 打开配置对话框失败: {e}")
                    import traceback
                    traceback.print_exc()
            
            def download_txt(self):
                """下载表格为TXT文本（表格形式）"""
                try:
                    # 收集当前数据
                    self.collect_data()
                    
                    # 构建TXT内容（表格形式，使用制表符分隔）
                    txt_content = ""
                    
                    # 表头行
                    txt_content += "\t".join(self.headers) + "\n"
                    
                    # 数据行
                    for idx, row in enumerate(self.table_data):
                        # 镜号
                        row_data = [str(idx + 1)]
                        # 其他列
                        row_data.extend([str(cell) for cell in row])
                        txt_content += "\t".join(row_data) + "\n"
                    
                    # 保存文件
                    from PySide6.QtWidgets import QFileDialog
                    file_path, _ = QFileDialog.getSaveFileName(
                        None,
                        "保存TXT文件",
                        f"{self.original_table_name}.txt",
                        "文本文件 (*.txt);;所有文件 (*.*)"
                    )
                    
                    if file_path:
                        with open(file_path, 'w', encoding='utf-8') as f:
                            f.write(txt_content)
                        
                        print(f"[表格节点] 已下载为TXT: {file_path}")
                        print(f"[表格节点] 共 {len(self.table_data)} 行数据")
                
                except Exception as e:
                    print(f"[表格节点] 下载TXT失败: {e}")
                    import traceback
                    traceback.print_exc()
            
            def upload_txt(self):
                """上传TXT文本并导入到表格"""
                try:
                    from PySide6.QtWidgets import QFileDialog, QMessageBox
                    
                    # 选择文件
                    file_path, _ = QFileDialog.getOpenFileName(
                        None,
                        "选择TXT文件",
                        "",
                        "文本文件 (*.txt);;所有文件 (*.*)"
                    )
                    
                    if not file_path:
                        return
                    
                    # 读取文件
                    with open(file_path, 'r', encoding='utf-8') as f:
                        lines = f.readlines()
                    
                    if len(lines) < 2:
                        QMessageBox.warning(None, "错误", "TXT文件格式不正确，至少需要表头和一行数据")
                        return
                    
                    # 解析表头
                    header_line = lines[0].strip()
                    new_headers = [h.strip() for h in header_line.split('\t')]
                    
                    # 解析数据
                    new_data = []
                    for line in lines[1:]:
                        line = line.strip()
                        if not line:
                            continue
                        
                        cells = line.split('\t')
                        # 跳过第一列（镜号，会自动生成）
                        if len(cells) > 1:
                            new_data.append(cells[1:])
                    
                    # 确认导入
                    reply = QMessageBox.question(
                        None,
                        "确认导入",
                        f"将导入 {len(new_data)} 行数据，{len(new_headers)} 列\n\n是否继续？",
                        QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
                    )
                    
                    if reply == QMessageBox.StandardButton.Yes:
                        # 更新表头和数据
                        self.headers = new_headers
                        self.table_data = new_data
                        
                        # 重新计算列宽
                        num_cols = len(new_headers)
                        if num_cols <= 5:
                            self.column_widths = [60] + [150] * (num_cols - 1)
                        else:
                            self.column_widths = [60] + [120] * (num_cols - 1)
                        
                        # 刷新表格
                        self.refresh_table()
                        
                        print(f"[表格节点] 已导入TXT: {file_path}")
                        print(f"[表格节点] 表头: {new_headers}")
                        print(f"[表格节点] 数据行数: {len(new_data)}")
                
                except Exception as e:
                    print(f"[表格节点] 上传TXT失败: {e}")
                    import traceback
                    traceback.print_exc()
                    QMessageBox.critical(None, "错误", f"导入失败: {str(e)}")
                    QMessageBox.warning(None, "错误", f"上传文本失败:\n{str(e)}")
            
            def download_excel(self):
                """下载为Excel文件"""
                try:
                    # 收集当前数据
                    self.collect_data()
                    
                    # 弹出文件保存对话框
                    file_path, _ = QFileDialog.getSaveFileName(
                        None,
                        "保存Excel文件",
                        f"{self.original_table_name}.xlsx",
                        "Excel文件 (*.xlsx)"
                    )
                    
                    if not file_path:
                        return
                    
                    # 导入openpyxl（如果未安装会提示）
                    try:
                        from openpyxl import Workbook
                        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
                    except ImportError:
                        QMessageBox.warning(
                            None,
                            "缺少依赖",
                            "需要安装 openpyxl 库才能导出Excel文件。\n\n请在终端运行：\npip install openpyxl"
                        )
                        return
                    
                    # 创建工作簿
                    wb = Workbook()
                    ws = wb.active
                    ws.title = self.original_table_name[:31]  # Excel工作表名称限制31字符
                    
                    # 定义样式
                    header_font = Font(name='Microsoft YaHei', size=11, bold=True, color='FFFFFF')
                    header_fill = PatternFill(start_color='667EEA', end_color='764BA2', fill_type='solid')
                    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    
                    cell_font = Font(name='Microsoft YaHei', size=10)
                    cell_alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
                    cell_alignment_center = Alignment(horizontal='center', vertical='center')
                    
                    border = Border(
                        left=Side(style='thin', color='D0D0D0'),
                        right=Side(style='thin', color='D0D0D0'),
                        top=Side(style='thin', color='D0D0D0'),
                        bottom=Side(style='thin', color='D0D0D0')
                    )
                    
                    # 写入表头
                    for col_idx, header in enumerate(self.headers, start=1):
                        cell = ws.cell(row=1, column=col_idx, value=header)
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.alignment = header_alignment
                        cell.border = border
                    
                    # 写入数据行
                    for row_idx, row_data in enumerate(self.table_data, start=2):
                        # 第一列：镜号
                        cell = ws.cell(row=row_idx, column=1, value=row_idx - 1)
                        cell.font = Font(name='Microsoft YaHei', size=10, bold=True, color='667EEA')
                        cell.alignment = cell_alignment_center
                        cell.border = border
                        cell.fill = PatternFill(start_color='F5F7FF', end_color='F5F7FF', fill_type='solid')
                        
                        # 其他列：数据
                        for col_idx, cell_value in enumerate(row_data, start=2):
                            cell = ws.cell(row=row_idx, column=col_idx, value=str(cell_value))
                            cell.font = cell_font
                            cell.alignment = cell_alignment
                            cell.border = border
                            
                            # 偶数行添加背景色
                            if row_idx % 2 == 0:
                                cell.fill = PatternFill(start_color='F9F9F9', end_color='F9F9F9', fill_type='solid')
                    
                    # 设置列宽
                    column_widths = {
                        1: 8,   # 镜号
                        2: 12,  # 景别
                        3: 30,  # 画面内容
                        4: 15,  # 人物
                        5: 18,  # 人物关系/构图
                        6: 15,  # 地点/环境
                        7: 12,  # 运镜
                        8: 25,  # 音效/台词
                        9: 15,  # 备注
                    }
                    
                    # 如果有额外的列（如提示词列），设置更大的宽度
                    for col_idx in range(1, len(self.headers) + 1):
                        if col_idx <= len(column_widths):
                            ws.column_dimensions[chr(64 + col_idx)].width = column_widths.get(col_idx, 15)
                        else:
                            # 提示词列等额外列使用更大宽度
                            ws.column_dimensions[chr(64 + col_idx)].width = 35
                    
                    # 设置行高
                    ws.row_dimensions[1].height = 25  # 表头行高
                    for row_idx in range(2, len(self.table_data) + 2):
                        ws.row_dimensions[row_idx].height = 60  # 数据行高
                    
                    # 保存文件
                    wb.save(file_path)
                    
                    print(f"[表格节点] Excel文件已保存: {file_path}")
                    QMessageBox.information(None, "导出成功", f"Excel文件已保存到：\n{file_path}")
                    
                except Exception as e:
                    print(f"[表格节点] 导出Excel失败: {e}")
                    import traceback
                    traceback.print_exc()
            
            def open_settings_dialog(self):
                """打开配置对话框"""
                try:
                    from lingdongsetting import StoryboardSettingsDialog
                    
                    # 当前配置
                    current_settings = {
                        'cell_height': self.cell_height,
                        'header_height': self.header_height,
                        'auto_number': True,
                        'highlight_color': '#ff0000',
                        'text_color': '#cccccc',
                        'header_color': '#00bfff',
                        'show_grid': True,
                        'max_text_length': 80
                    }
                    
                    # 创建并显示对话框
                    dialog = StoryboardSettingsDialog(None, current_settings)
                    
                    if dialog.exec():
                        # 应用新配置
                        new_settings = dialog.get_settings()
                        self.cell_height = new_settings['cell_height']
                        self.header_height = new_settings['header_height']
                        
                        # 刷新表格
                        self.refresh_table()
                        
                        print(f"[表格配置] 已更新: 单元格高度={self.cell_height}, 表头高度={self.header_height}")
                
                except Exception as e:
                    print(f"[表格配置] 打开配置对话框失败: {e}")
                    import traceback
                    traceback.print_exc()
            
            def download_txt(self):
                """下载表格为TXT文本（表格形式）"""
                try:
                    # 收集当前数据
                    self.collect_data()
                    
                    # 构建TXT内容（表格形式，使用制表符分隔）
                    txt_content = ""
                    
                    # 表头行
                    txt_content += "\t".join(self.headers) + "\n"
                    
                    # 数据行
                    for idx, row in enumerate(self.table_data):
                        # 镜号
                        row_data = [str(idx + 1)]
                        # 其他列
                        row_data.extend([str(cell) for cell in row])
                        txt_content += "\t".join(row_data) + "\n"
                    
                    # 保存文件
                    from PySide6.QtWidgets import QFileDialog
                    file_path, _ = QFileDialog.getSaveFileName(
                        None,
                        "保存TXT文件",
                        f"{self.original_table_name}.txt",
                        "文本文件 (*.txt);;所有文件 (*.*)"
                    )
                    
                    if file_path:
                        with open(file_path, 'w', encoding='utf-8') as f:
                            f.write(txt_content)
                        
                        print(f"[表格节点] 已下载为TXT: {file_path}")
                        print(f"[表格节点] 共 {len(self.table_data)} 行数据")
                
                except Exception as e:
                    print(f"[表格节点] 下载TXT失败: {e}")
                    import traceback
                    traceback.print_exc()
            
            def upload_txt(self):
                """上传TXT文本并导入到表格"""
                try:
                    from PySide6.QtWidgets import QFileDialog, QMessageBox
                    
                    # 选择文件
                    file_path, _ = QFileDialog.getOpenFileName(
                        None,
                        "选择TXT文件",
                        "",
                        "文本文件 (*.txt);;所有文件 (*.*)"
                    )
                    
                    if not file_path:
                        return
                    
                    # 读取文件
                    with open(file_path, 'r', encoding='utf-8') as f:
                        lines = f.readlines()
                    
                    if len(lines) < 2:
                        QMessageBox.warning(None, "错误", "TXT文件格式不正确，至少需要表头和一行数据")
                        return
                    
                    # 解析表头
                    header_line = lines[0].strip()
                    new_headers = [h.strip() for h in header_line.split('\t')]
                    
                    # 解析数据
                    new_data = []
                    for line in lines[1:]:
                        line = line.strip()
                        if not line:
                            continue
                        
                        cells = line.split('\t')
                        # 跳过第一列（镜号，会自动生成）
                        if len(cells) > 1:
                            new_data.append(cells[1:])
                    
                    # 确认导入
                    reply = QMessageBox.question(
                        None,
                        "确认导入",
                        f"将导入 {len(new_data)} 行数据，{len(new_headers)} 列\n\n是否继续？",
                        QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
                    )
                    
                    if reply == QMessageBox.StandardButton.Yes:
                        # 更新表头和数据
                        self.headers = new_headers
                        self.table_data = new_data
                        
                        # 重新计算列宽
                        num_cols = len(new_headers)
                        if num_cols <= 5:
                            self.column_widths = [60] + [150] * (num_cols - 1)
                        else:
                            self.column_widths = [60] + [120] * (num_cols - 1)
                        
                        # 刷新表格
                        self.refresh_table()
                        
                        print(f"[表格节点] 已导入TXT: {file_path}")
                        print(f"[表格节点] 表头: {new_headers}")
                        print(f"[表格节点] 数据行数: {len(new_data)}")
                
                except Exception as e:
                    print(f"[表格节点] 上传TXT失败: {e}")
                    import traceback
                    traceback.print_exc()
                    QMessageBox.critical(None, "错误", f"导入失败: {str(e)}")
                    QMessageBox.warning(None, "错误", f"导出Excel失败:\n{str(e)}")
        
        return StoryboardNodeImpl
